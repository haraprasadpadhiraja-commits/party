"""
Application views for the Fresher Party Management System.

Public views:      home, register, success, faculty, search, gallery, about
AJAX live search:  live_search
Admin views:       login (Django built-in), dashboard, students list, add/edit/
                   delete student, faculty management, college edit, gallery
                   management, exports (Excel / PDF)
"""

import io

from django.contrib import messages
from django.contrib.auth import login as auth_login
from django.contrib.auth.decorators import login_required
from django.contrib.auth.forms import AuthenticationForm
from django.db.models import Count, Q
from django.http import HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse
from django.utils import timezone
from django.views.decorators.http import require_POST

# --- Excel export ---
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

# --- PDF export ---
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.units import inch
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import (
    SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle,
)

from .forms import (
    CollegeInfoForm, FacultyForm, GalleryImageForm,
    StudentForm, StudentRegistrationForm,
)
from .models import CollegeInfo, Faculty, GalleryImage, Student


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def get_college_info():
    return CollegeInfo.get_info()


def page_context(request, **extra):
    """Shared context used on every page."""
    ctx = {
        'college': get_college_info(),
        'student_count': Student.objects.count(),
        'faculty_count': Faculty.objects.count(),
        'branch_count': len(Student.objects.values_list('branch', flat=True).distinct()),
    }
    ctx.update(extra)
    return ctx


def _search_students(q):
    """Return students matching the query across common fields."""
    if not q:
        return Student.objects.none()
    qs = Student.objects.filter(
        Q(name__icontains=q) | Q(registration_number__icontains=q)
        | Q(branch__icontains=q) | Q(father_name__icontains=q)
        | Q(phone__icontains=q)
    )
    return qs


def _search_faculty(q):
    if not q:
        return Faculty.objects.none()
    return Faculty.objects.filter(
        Q(name__icontains=q) | Q(faculty_number__icontains=q)
        | Q(branch__icontains=q) | Q(designation__icontains=q)
    )


# ---------------------------------------------------------------------------
# Public pages
# ---------------------------------------------------------------------------

def home(request):
    ctx = page_context(request)
    ctx['students'] = Student.objects.order_by('-created_at')[:8]
    return render(request, 'home.html', ctx)


def student_register(request):
    form = StudentRegistrationForm(request.POST or None)
    if request.method == 'POST' and form.is_valid():
        student = form.save(commit=False)
        # Auto-generate a unique registration number (handles duplicates).
        student.registration_number = Student.generate_registration_number()
        while Student.objects.filter(
                registration_number=student.registration_number).exists():
            student.registration_number = Student.generate_registration_number()
        student.save()
        return redirect('registration_success', pk=student.pk)
    return render(request, 'student_register.html',
                  page_context(request, form=form))


def registration_success(request, pk):
    student = get_object_or_404(Student, pk=pk)
    return render(request, 'registration_success.html',
                  page_context(request, student=student))


def faculty_page(request):
    branch = request.GET.get('branch', '')
    q = request.GET.get('q', '').strip()
    qs = Faculty.objects.all()
    if branch:
        qs = qs.filter(branch__icontains=branch)
    if q:
        qs = _search_faculty(q)
    branches = sorted(Faculty.objects.values_list('branch', flat=True).distinct())
    return render(request, 'faculty.html',
                  page_context(request, faculty=qs, branches=branches,
                               current_branch=branch, q=q))


def search_page(request):
    q = request.GET.get('q', '').strip()
    students = _search_students(q)
    faculty = _search_faculty(q)
    return render(request, 'search.html',
                  page_context(request, q=q, students=students, faculty=faculty))


def live_search(request):
    """AJAX live search used by the search box in the navbar."""
    q = request.GET.get('q', '').strip()
    students = list(_search_students(q).values('id', 'registration_number',
                                               'name', 'branch')[:10])
    faculty = list(_search_faculty(q).values('id', 'faculty_number', 'name',
                                             'branch', 'designation')[:10])
    return JsonResponse({'query': q, 'students': students, 'faculty': faculty})


def gallery_page(request):
    category = request.GET.get('category', '')
    qs = GalleryImage.objects.all()
    if category:
        qs = qs.filter(category=category)
    categories = [c[0] for c in GalleryImage.CATEGORY_CHOICES]
    return render(request, 'gallery.html',
                  page_context(request, images=qs, categories=categories,
                               current_category=category))


def about_page(request):
    info = get_college_info()
    # Split newline-separated text into lists for nice bullets.
    def lines(text):
        return [x.strip() for x in (text or '').splitlines() if x.strip()]
    return render(request, 'about.html',
                  page_context(request, info=info,
                               courses=lines(info.courses),
                               branches=lines(info.branches),
                               facilities=lines(info.facilities)))


def dashboard(request):
    students = Student.objects.all()
    faculty = Faculty.objects.all()

    students_by_branch = list(
        Student.objects.values('branch').annotate(count=Count('id'))
        .order_by('-count'))
    faculty_by_branch = list(
        Faculty.objects.values('branch').annotate(count=Count('id'))
        .order_by('-count'))

    # Registration growth: students registered in the last 7 days (incl. today).
    today = timezone.now().date()
    days = []
    for offset in range(6, -1, -1):
        day = today - timezone.timedelta(days=offset)
        day_start = timezone.make_aware(
            timezone.datetime(day.year, day.month, day.day))
        day_end = day_start + timezone.timedelta(days=1)
        days.append({'label': day.strftime('%a'), 'count':
                     Student.objects.filter(created_at__gte=day_start,
                                            created_at__lt=day_end).count()})

    # Marks distribution buckets.
    marks_buckets = []
    for lo, hi in [(0, 40), (40, 60), (60, 75), (75, 90), (90, 101)]:
        marks_buckets.append({
            'label': f"{lo}-{hi - 1}" if hi <= 100 else f"{lo}-100",
            'count': Student.objects.filter(
                marks__gte=lo, marks__lt=hi if hi <= 100 else 101).count(),
        })

    return render(request, 'dashboard.html', page_context(request,
        students=students, faculty=faculty,
        students_by_branch=students_by_branch,
        faculty_by_branch=faculty_by_branch,
        growth_days=days, marks_buckets=marks_buckets))


# ---------------------------------------------------------------------------
# Admin views
# ---------------------------------------------------------------------------

def login_view(request):
    if request.user.is_authenticated:
        return redirect('admin_dashboard')
    form = AuthenticationForm(request, data=request.POST or None)
    if request.method == 'POST' and form.is_valid():
        auth_login(request, form.get_user())
        messages.success(request, f"Welcome back, {request.user.username}!")
        return redirect('admin_dashboard')
    return render(request, 'admin/login.html', {'form': form})


@login_required
def admin_dashboard(request):
    return render(request, 'admin/dashboard.html', page_context(request))


# --- Students ---

@login_required
def student_list(request):
    q = request.GET.get('q', '').strip()
    branch = request.GET.get('branch', '')
    sort = request.GET.get('sort', 'registration_number')
    qs = Student.objects.all()
    if q:
        qs = _search_students(q)
    if branch:
        qs = qs.filter(branch=branch)
    allowed = {
        'registration_number': 'registration_number',
        'name': 'name', 'branch': 'branch', 'marks': '-marks',
        'created_at': '-created_at',
    }
    if sort in allowed:
        qs = qs.order_by(allowed[sort])
    branches = [c[0] for c in Student.BRANCH_CHOICES]
    return render(request, 'admin/student_list.html',
                  page_context(request, students=qs, branches=branches,
                               q=q, current_branch=branch, sort=sort))


@login_required
def student_add(request):
    form = StudentForm(request.POST or None)
    if request.method == 'POST' and form.is_valid():
        student = form.save(commit=False)
        student.registration_number = Student.generate_registration_number()
        while Student.objects.filter(
                registration_number=student.registration_number).exists():
            student.registration_number = Student.generate_registration_number()
        student.save()
        messages.success(request, f"Student {student.name} added successfully.")
        return redirect('student_list')
    return render(request, 'admin/student_form.html',
                  page_context(request, form=form, title='Add Student',
                               action_url=reverse('student_add')))


@login_required
def student_edit(request, pk):
    student = get_object_or_404(Student, pk=pk)
    form = StudentForm(request.POST or None, instance=student)
    if request.method == 'POST' and form.is_valid():
        form.save()
        messages.success(request, f"Student {student.name} updated successfully.")
        return redirect('student_list')
    return render(request, 'admin/student_form.html',
                  page_context(request, form=form, title=f'Edit {student.name}',
                               action_url=reverse('student_edit', args=[pk])))


@login_required
def student_view(request, pk):
    student = get_object_or_404(Student, pk=pk)
    return render(request, 'admin/student_view.html',
                  page_context(request, student=student))


@login_required
@require_POST
def student_delete(request, pk):
    student = get_object_or_404(Student, pk=pk)
    student.delete()
    messages.success(request, f"Student {student.name} deleted successfully.")
    return redirect('student_list')


# --- Faculty ---

@login_required
def faculty_list(request):
    q = request.GET.get('q', '').strip()
    qs = Faculty.objects.all()
    if q:
        qs = _search_faculty(q)
    return render(request, 'admin/faculty_list.html',
                  page_context(request, faculty=qs, q=q))


@login_required
def faculty_add(request):
    form = FacultyForm(request.POST or None)
    if request.method == 'POST' and form.is_valid():
        member = form.save(commit=False)
        member.faculty_number = Faculty.generate_faculty_number()
        while Faculty.objects.filter(
                faculty_number=member.faculty_number).exists():
            member.faculty_number = Faculty.generate_faculty_number()
        member.save()
        messages.success(request, f"Faculty {member.name} added successfully.")
        return redirect('faculty_list')
    return render(request, 'admin/faculty_form.html',
                  page_context(request, form=form, title='Add Faculty',
                               action_url=reverse('faculty_add')))


@login_required
def faculty_edit(request, pk):
    member = get_object_or_404(Faculty, pk=pk)
    form = FacultyForm(request.POST or None, instance=member)
    if request.method == 'POST' and form.is_valid():
        form.save()
        messages.success(request, f"Faculty {member.name} updated successfully.")
        return redirect('faculty_list')
    return render(request, 'admin/faculty_form.html',
                  page_context(request, form=form, title=f'Edit {member.name}',
                               action_url=reverse('faculty_edit', args=[pk])))


@login_required
@require_POST
def faculty_delete(request, pk):
    member = get_object_or_404(Faculty, pk=pk)
    member.delete()
    messages.success(request, f"Faculty {member.name} deleted successfully.")
    return redirect('faculty_list')


# --- College information ---

@login_required
def college_edit(request):
    info = get_college_info()
    form = CollegeInfoForm(request.POST or None, request.FILES or None,
                           instance=info)
    if request.method == 'POST' and form.is_valid():
        form.save()
        messages.success(request, "College information updated successfully.")
        return redirect('about_page')
    return render(request, 'admin/college_edit.html',
                  page_context(request, form=form))


# --- Gallery management ---

@login_required
def gallery_manage(request):
    images = GalleryImage.objects.all()
    return render(request, 'admin/gallery_manage.html',
                  page_context(request, images=images))


@login_required
def gallery_add(request):
    form = GalleryImageForm(request.POST or None, request.FILES or None)
    if request.method == 'POST' and form.is_valid():
        form.save()
        messages.success(request, "Gallery image added successfully.")
        return redirect('gallery_manage')
    return render(request, 'admin/gallery_form.html',
                  page_context(request, form=form, title='Add Gallery Image',
                               action_url=reverse('gallery_add')))


@login_required
def gallery_edit(request, pk):
    image = get_object_or_404(GalleryImage, pk=pk)
    form = GalleryImageForm(request.POST or None, request.FILES or None,
                            instance=image)
    if request.method == 'POST' and form.is_valid():
        form.save()
        messages.success(request, "Gallery image updated successfully.")
        return redirect('gallery_manage')
    return render(request, 'admin/gallery_form.html',
                  page_context(request, form=form, title=f'Edit {image.title}',
                               action_url=reverse('gallery_edit', args=[pk])))


@login_required
@require_POST
def gallery_delete(request, pk):
    image = get_object_or_404(GalleryImage, pk=pk)
    image.delete()
    messages.success(request, "Gallery image deleted successfully.")
    return redirect('gallery_manage')


# ---------------------------------------------------------------------------
# Export functionality
# ---------------------------------------------------------------------------

def _student_rows(students):
    """Shared list of columns for Excel and PDF exports."""
    return [
        ['Registration No.', 'Student Name', "Father's Name", 'Address',
         'Phone', 'Marks', 'Branch', 'Registration Date'],
    ] + [
        [s.registration_number, s.name, s.father_name, s.address, s.phone,
         str(s.marks), s.branch,
         timezone.localtime(s.created_at).strftime('%d-%b-%Y %H:%M')]
        for s in students
    ]


@login_required
def export_excel(request):
    """Export students to an .xlsx file with styled headers."""
    q = request.GET.get('q', '').strip()
    students = _search_students(q) if q else Student.objects.all()
    rows = _student_rows(students)

    wb = Workbook()
    ws = wb.active
    ws.title = 'Students'

    header_fill = PatternFill('solid', fgColor='4F46E5')
    header_font = Font(color='FFFFFF', bold=True, size=12)

    for col_idx, cell in enumerate(rows[0], start=1):
        c = ws.cell(row=1, column=col_idx, value=cell)
        c.font = header_font
        c.fill = header_fill
        c.alignment = Alignment(horizontal='center')

    for r in rows[1:]:
        ws.append(r)

    # Column widths & freeze header row.
    widths = [18, 24, 20, 30, 14, 10, 22, 18]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = 'A2'

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    response = HttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="students_export.xlsx"'
    return response


@login_required
def export_pdf(request):
    """Export students to a professional PDF with a college header."""
    info = get_college_info()
    q = request.GET.get('q', '').strip()
    students = _search_students(q) if q else Student.objects.all()
    rows = _student_rows(students)

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4),
                            rightMargin=0.5 * inch, leftMargin=0.5 * inch,
                            topMargin=0.6 * inch, bottomMargin=0.6 * inch)
    styles = getSampleStyleSheet()

    title_style = ParagraphStyle(
        'TitleStyle', parent=styles['Title'], fontSize=16, alignment=1,
        textColor=colors.HexColor('#4F46E5'), spaceAfter=2)
    table_header_style = ParagraphStyle(
        'TableHeader', parent=styles['Normal'], fontSize=8,
        textColor=colors.white, fontName='Helvetica-Bold')
    sub_style = ParagraphStyle(
        'SubStyle', parent=styles['Normal'], alignment=1, fontSize=10,
        textColor=colors.HexColor('#555555'), spaceAfter=6)

    header = [
        Paragraph(info.college_name, title_style),
        Paragraph(f"Fresher Party Registration List — {timezone.now().year}",
                  sub_style),
        Paragraph(f"Total Students: {students.count()}   |   "
                  f"Generated: {timezone.localtime(timezone.now()).strftime('%d %b %Y %H:%M')}",
                  sub_style),
        Spacer(1, 6),
    ]

    data = [[Paragraph(str(c), table_header_style) for c in rows[0]]] + rows[1:]
    table = Table(data, repeatRows=1)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4F46E5')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 8),
        ('FONTSIZE', (0, 1), (-1, -1), 7),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1),
         [colors.white, colors.HexColor('#EEF2FF')]),
        ('GRID', (0, 0), (-1, -1), 0.4, colors.HexColor('#C7D2FE')),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('LEFTPADDING', (0, 0), (-1, -1), 4),
        ('RIGHTPADDING', (0, 0), (-1, -1), 4),
    ]))
    header.append(table)

    doc.build(header)
    buffer.seek(0)

    response = HttpResponse(buffer.getvalue(),
                            content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="students_export.pdf"'
    return response
